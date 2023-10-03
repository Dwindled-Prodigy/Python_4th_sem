from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title="Capital")
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']
headers = ["State", "Language", "Code"]
for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]
wb.save("demo.xlsx")
sheet = wb["Capital"]
for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = capital[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]
wb.save("demo.xlsx")
sheet = wb["Capital"]
srchCode = input("Enter state code for finding capital: ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)
sheet = wb["Language"]
srchCode = input("Enter state code for finding language: ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row=i, column=2).value)
wb.close()